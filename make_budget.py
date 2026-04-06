from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── COLOURS ──────────────────────────────────────────────────────────────────
C_DARK   = "1A1C1A"
C_BROWN  = "8B4F2D"
C_LIGHT  = "F4F3F1"
C_WHITE  = "FFFFFF"
C_GREEN  = "D6EAD6"
C_RED    = "FAD6D6"
C_YELLOW = "FFF3CD"

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(bold=False, color=C_DARK, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def border_bottom():
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)
def usd(ws, row, col):
    ws.cell(row, col).number_format = '"$"#,##0.00'
def pct(ws, row, col):
    ws.cell(row, col).number_format = '0.0"%"'

def header_row(ws, row, texts, widths=None, bg=C_DARK, fg=C_WHITE):
    for i, t in enumerate(texts, 1):
        c = ws.cell(row, i, t)
        c.font = font(bold=True, color=fg, size=11)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

def section_title(ws, row, text):
    c = ws.cell(row, 1, text)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_BROWN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def data_row(ws, row, label, vals, fmt="usd", shade=False):
    c = ws.cell(row, 1, label)
    c.font = font(size=10)
    c.fill = fill(C_LIGHT) if shade else fill(C_WHITE)
    c.alignment = Alignment(indent=2)
    for i, v in enumerate(vals, 2):
        cell = ws.cell(row, i, v)
        cell.font = font(size=10)
        cell.fill = fill(C_LIGHT) if shade else fill(C_WHITE)
        cell.alignment = Alignment(horizontal="right")
        if fmt == "usd":
            cell.number_format = '"$"#,##0'
        elif fmt == "pct":
            cell.number_format = '0.0"%"'
        elif fmt == "int":
            cell.number_format = '#,##0'

def total_row(ws, row, label, formulas, bg=C_DARK, fg=C_WHITE, fmt="usd"):
    c = ws.cell(row, 1, label)
    c.font = font(bold=True, color=fg, size=10)
    c.fill = fill(bg)
    c.alignment = Alignment(indent=1)
    for i, f in enumerate(formulas, 2):
        cell = ws.cell(row, i, f)
        cell.font = font(bold=True, color=fg, size=10)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="right")
        if fmt == "usd":
            cell.number_format = '"$"#,##0'
        elif fmt == "pct":
            cell.number_format = '0.0"%"'

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — VARSAYIMLAR (Inputs)
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Varsayımlar"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.column_dimensions["A"].width = 34
for col in ["B","C","D"]:
    ws1.column_dimensions[col].width = 16

# Title
t = ws1.cell(1, 1, "PALLET Leather — Model Varsayımları")
t.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
t.fill = fill(C_DARK)
t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws1.merge_cells("A1:D1")

header_row(ws1, 2, ["Parametre", "Yıl 1", "Yıl 2", "Yıl 3"],
           widths=[34,16,16,16], bg=C_BROWN)

# Satış hacmi
section_title(ws1, 3, "SATIŞ HACMİ")
data_row(ws1, 4,  "Yıllık satış (çift)", [180, 360, 600], fmt="int")
data_row(ws1, 5,  "Aylık ortalama (çift)", [15, 30, 50], fmt="int", shade=True)

# Fiyatlandırma
section_title(ws1, 6, "FİYATLANDIRMA")
data_row(ws1, 7,  "Satış fiyatı / çift ($)", [35, 40, 45])
data_row(ws1, 8,  "Birim üretim maliyeti ($)", [15, 15, 15], shade=True)
data_row(ws1, 9,  "Birim brüt kar ($)", ["=C7-C8","=D7-D8","=E7-E8"])

# Pazarlama
section_title(ws1, 10, "PAZARLAMA")
data_row(ws1, 11, "Aylık pazarlama bütçesi ($)", [500, 700, 1000])
data_row(ws1, 12, "Yıllık pazarlama bütçesi ($)", ["=C11*12","=D11*12","=E11*12"], shade=True)

# Sabit giderler
section_title(ws1, 13, "SABİT GİDERLER (AYLIK)")
data_row(ws1, 14, "Depo / Kargo / Paketleme ($)", [250, 250, 250])
data_row(ws1, 15, "Web, domain, araçlar ($)", [50, 50, 100], shade=True)
data_row(ws1, 16, "Muhasebe & diğer ($)", [100, 100, 100])

# note
ws1.cell(18, 1, "💡 Sarı hücrelerdeki değerleri değiştirerek tüm model otomatik güncellenir.").font = Font(
    italic=True, color="888888", size=9, name="Calibri")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — P&L
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📊 P&L")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 34
for col in ["B","C","D","E"]:
    ws2.column_dimensions[col].width = 16

# Title
t = ws2.cell(1, 1, "PALLET Leather — Gelir Tablosu (3 Yıl)")
t.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
t.fill = fill(C_DARK)
t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws2.merge_cells("A1:E1")
ws2.row_dimensions[1].height = 40

header_row(ws2, 2, ["", "Yıl 1", "Yıl 2", "Yıl 3", "3 Yıl Toplam"],
           widths=[34,16,16,16,16], bg=C_BROWN)

# GELİR
section_title(ws2, 3, "GELİR")
# Refs to varsayimlar sheet
data_row(ws2, 4, "Satış adedi (çift)", [
    "='📋 Varsayımlar'!C4",
    "='📋 Varsayımlar'!D4",
    "='📋 Varsayımlar'!E4",
    "=B4+C4+D4"
], fmt="int")
data_row(ws2, 5, "Ortalama satış fiyatı ($)", [
    "='📋 Varsayımlar'!C7",
    "='📋 Varsayımlar'!D7",
    "='📋 Varsayımlar'!E7",
    "=AVERAGE(B5:D5)"
], shade=True)
total_row(ws2, 6, "TOPLAM GELİR", [
    "=B4*B5", "=C4*C5", "=D4*D5", "=B6+C6+D6"
], bg=C_BROWN)

# COGS
section_title(ws2, 7, "SATILAN MAL MALİYETİ (COGS)")
data_row(ws2, 8, "Birim üretim maliyeti ($)", [
    "='📋 Varsayımlar'!C8",
    "='📋 Varsayımlar'!D8",
    "='📋 Varsayımlar'!E8",
    "=AVERAGE(B8:D8)"
])
total_row(ws2, 9, "TOPLAM COGS", [
    "=B4*B8", "=C4*C8", "=D4*D8", "=B9+C9+D9"
], bg="555555")

# BRÜT KAR
total_row(ws2, 10, "BRÜT KAR", [
    "=B6-B9", "=C6-C9", "=D6-D9", "=B10+C10+D10"
], bg=C_DARK)

data_row(ws2, 11, "  Brüt Marj (%)", [
    '=IF(B6>0,B10/B6*100,0)',
    '=IF(C6>0,C10/C6*100,0)',
    '=IF(D6>0,D10/D6*100,0)',
    '=IF(E6>0,E10/E6*100,0)'
], fmt="pct", shade=True)

# GİDERLER
section_title(ws2, 12, "OPERASYONEL GİDERLER")
data_row(ws2, 13, "Pazarlama", [
    "='📋 Varsayımlar'!C12",
    "='📋 Varsayımlar'!D12",
    "='📋 Varsayımlar'!E12",
    "=B13+C13+D13"
])
data_row(ws2, 14, "Depo / Kargo / Paketleme", [
    "='📋 Varsayımlar'!C14*12",
    "='📋 Varsayımlar'!D14*12",
    "='📋 Varsayımlar'!E14*12",
    "=B14+C14+D14"
], shade=True)
data_row(ws2, 15, "Web, domain, araçlar", [
    "='📋 Varsayımlar'!C15*12",
    "='📋 Varsayımlar'!D15*12",
    "='📋 Varsayımlar'!E15*12",
    "=B15+C15+D15"
])
data_row(ws2, 16, "Muhasebe & diğer", [
    "='📋 Varsayımlar'!C16*12",
    "='📋 Varsayımlar'!D16*12",
    "='📋 Varsayımlar'!E16*12",
    "=B16+C16+D16"
], shade=True)
total_row(ws2, 17, "TOPLAM GİDER", [
    "=B13+B14+B15+B16",
    "=C13+C14+C15+C16",
    "=D13+D14+D15+D16",
    "=E13+E14+E15+E16"
], bg="555555")

# NET KAR
ws2.row_dimensions[18].height = 24
for col in range(1, 6):
    cell = ws2.cell(18, col)
    cell.fill = fill(C_DARK)
    cell.font = font(bold=True, color=C_WHITE, size=12)
    cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center", indent=1 if col == 1 else 0)
    cell.number_format = '"$"#,##0'

ws2.cell(18, 1, "NET KAR / ZARAR")
for i, f in enumerate(["=B10-B17","=C10-C17","=D10-D17","=E10-E17"], 2):
    c = ws2.cell(18, i, f)
    c.number_format = '"$"#,##0'
    c.font = font(bold=True, color=C_WHITE, size=12)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal="right")

# Conditional colour for net profit row (manual — green/red)
data_row(ws2, 19, "  Net Marj (%)", [
    '=IF(B6>0,B18/B6*100,0)',
    '=IF(C6>0,C18/C6*100,0)',
    '=IF(D6>0,D18/D6*100,0)',
    '=IF(E6>0,E18/E6*100,0)'
], fmt="pct", shade=True)

# KÜMÜLATİF
section_title(ws2, 20, "KÜMÜLATİF NAKİT POZİSYONU")
data_row(ws2, 21, "Kümülatif Net", [
    "=B18", "=B21+C18", "=C21+D18", ""
])

# BAŞABAŞ
section_title(ws2, 22, "BAŞABAŞ ANALİZİ")
data_row(ws2, 23, "Başabaş (çift/ay)", [
    "=ROUND((B13/12+B14/12+B15/12+B16/12)/(B5-B8),0)",
    "=ROUND((C13/12+C14/12+C15/12+C16/12)/(C5-C8),0)",
    "=ROUND((D13/12+D14/12+D15/12+D16/12)/(D5-D8),0)",
    ""
], fmt="int")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Yıl 1 Aylık
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📅 Yıl 1 — Aylık")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 18
for i in range(2, 9):
    ws3.column_dimensions[get_column_letter(i)].width = 13

t = ws3.cell(1, 1, "PALLET Leather — Yıl 1 Aylık Döküm")
t.font = Font(bold=True, color=C_WHITE, size=14, name="Calibri")
t.fill = fill(C_DARK)
t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws3.merge_cells("A1:H1")
ws3.row_dimensions[1].height = 40

months = ["Ay 1","Ay 2","Ay 3","Ay 4","Ay 5","Ay 6","Ay 7","Ay 8","Ay 9","Ay 10","Ay 11","Ay 12","TOPLAM"]
sales  = [3,4,5,8,12,15,18,20,22,24,24,25]
price  = 35
cost   = 15
mkt_mo = 500
fixed  = 250+50+100  # 400/month

header_row(ws3, 2,
    ["", "Satış (çift)", "Gelir", "COGS", "Brüt Kar", "Gider", "Net", "Kümülatif"],
    widths=[18,13,13,13,13,13,13,13], bg=C_BROWN)

cum = 0
for i, mo in enumerate(months[:-1]):
    r = 3 + i
    s = sales[i]
    rev = s * price
    cogs = s * cost
    gp = rev - cogs
    opex = mkt_mo + fixed
    net = gp - opex
    cum += net
    shade = (i % 2 == 1)
    bg = C_LIGHT if shade else C_WHITE
    row_data = [mo, s, rev, cogs, gp, opex, net, cum]
    for j, v in enumerate(row_data, 1):
        c = ws3.cell(r, j, v)
        c.font = font(size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="right" if j > 1 else "left", indent=1 if j == 1 else 0)
        if j == 1:
            c.number_format = "@"
        elif j == 2:
            c.number_format = "#,##0"
        else:
            c.number_format = '"$"#,##0'
        # colour net and cumulative
        if j == 7:
            c.font = Font(bold=True, color=("006400" if v >= 0 else "8B0000"), size=10, name="Calibri")
        if j == 8:
            c.font = Font(bold=True, color=("006400" if v >= 0 else "8B0000"), size=10, name="Calibri")

# Totals row
r = 15
total_vals = ["TOPLAM", sum(sales), sum(s*price for s in sales),
    sum(s*cost for s in sales),
    sum(s*(price-cost) for s in sales),
    12*(mkt_mo+fixed),
    sum(s*(price-cost) for s in sales) - 12*(mkt_mo+fixed),
    ""]
for j, v in enumerate(total_vals, 1):
    c = ws3.cell(r, j, v)
    c.font = font(bold=True, color=C_WHITE, size=10)
    c.fill = fill(C_DARK)
    c.alignment = Alignment(horizontal="right" if j > 1 else "left", indent=1 if j == 1 else 0)
    if j >= 3 and isinstance(v, (int,float)):
        c.number_format = '"$"#,##0'
    elif j == 2:
        c.number_format = "#,##0"

output = "/home/user/palet/PALLET_Leather_Budget_Model.xlsx"
wb.save(output)
print(f"Saved: {output}")
